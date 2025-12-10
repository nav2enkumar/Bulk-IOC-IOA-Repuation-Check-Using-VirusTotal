import tkinter as tk
from tkinter import ttk, scrolledtext
from tkinter import messagebox
import threading
import requests
import pandas as pd
import base64
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

def virustotal_call(api_key, resource, resource_type):
    headers = {
        "x-apikey": api_key
    }
    url = f"https://www.virustotal.com/api/v3/{resource_type}/{resource}"
    response = requests.get(url, headers=headers)
    return response.json()

def process_md5(api_key, md5_hashes):
    md5_data = []
    for md5 in md5_hashes:
        response = virustotal_call(api_key, md5, "files")
        data = response.get('data', {})
        attributes = data.get('attributes', {})
        sha256 = attributes.get('sha256', 'N/A')
        last_analysis_stats = attributes.get('last_analysis_stats', {})
        total_votes = sum(last_analysis_stats.values())
        malicious_votes = last_analysis_stats.get('malicious', 0)
        detection_score = f"{malicious_votes} out of {total_votes}"
        
        action = "Block in EDR" if malicious_votes > 0 else "None"
        
        md5_data.append({
            'Category': 'Malware',
            'SHA-256': sha256,
            'VT Score': detection_score,
            'Action': action
        })
    return pd.DataFrame(md5_data)

def process_ips(api_key, ips):
    ip_data = []
    for ip in ips:
        response = virustotal_call(api_key, ip, "ip_addresses")
        data = response.get('data', {})
        attributes = data.get('attributes', {})
        last_analysis_stats = attributes.get('last_analysis_stats', {})
        total_votes = sum(last_analysis_stats.values())
        malicious_votes = last_analysis_stats.get('malicious', 0)
        detection_score = f"{malicious_votes} out of {total_votes}"
        as_owner = attributes.get('as_owner', 'N/A')
        country = attributes.get('country', 'N/A')
        
        action = "Block in F/W" if malicious_votes > 0 else "None"
        
        ip_data.append({
            'Category': 'IP',
            'IP Address': ip,
            'Organization': as_owner,
            'Country': country,
            'VT Score': detection_score,
            'Action': action
        })
    return pd.DataFrame(ip_data)

def process_urls(api_key, urls):
    url_data = []
    for url in urls:
        url_id = base64.urlsafe_b64encode(url.encode()).decode().strip("=")
        response = virustotal_call(api_key, url_id, "urls")
        data = response.get('data', {})
        attributes = data.get('attributes', {})
        last_analysis_stats = attributes.get('last_analysis_stats', {})
        total_votes = sum(last_analysis_stats.values())
        malicious_votes = last_analysis_stats.get('malicious', 0)
        detection_score = f"{malicious_votes} out of {total_votes}"
        
        action = "Block in F/W" if malicious_votes > 0 else "None"
        
        url_data.append({
            'Category': 'URL',
            'Address': url,
            'VT Score': detection_score,
            'Action': action
        })
    return pd.DataFrame(url_data)

def format_worksheet(worksheet, sheet_name):
    # Define styles
    header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_aligned = Alignment(horizontal="center", vertical="center")
    
    # Get the data range
    data_rows = worksheet.max_row
    data_cols = worksheet.max_column
    
    # Format headers and add borders
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="000000")
        cell.alignment = center_aligned
        cell.border = thin_border
    
    # Merge Category cells and format data
    first_category_cell = None
    last_category_cell = None
    category_value = None
    
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        current_category = row[0].value
        
        if category_value is None:
            category_value = current_category
            first_category_cell = row[0]
        elif row_idx == worksheet.max_row:
            # Last row - merge if needed
            if category_value == current_category:
                last_category_cell = row[0]
                if first_category_cell != last_category_cell:
                    merge_range = f"{first_category_cell.coordinate}:{last_category_cell.coordinate}"
                    worksheet.merge_cells(merge_range)
                    first_category_cell.alignment = center_aligned
                    
                    # Add borders to merged cells
                    for merge_row in range(first_category_cell.row, last_category_cell.row + 1):
                        worksheet.cell(row=merge_row, column=1).border = thin_border
        
        # Format data cells
        for cell in row:
            cell.alignment = center_aligned
            cell.border = thin_border
            
            # Apply red font for VT Score column if detections found
            if "VT Score" in worksheet.cell(row=1, column=cell.column).value:
                score = cell.value
                if score and "out of" in score:
                    detections = int(score.split()[0])
                    if detections > 0:
                        cell.font = Font(color="FF0000", bold=True)
            
            # Format Action column
            if "Action" in worksheet.cell(row=1, column=cell.column).value:
                if cell.value != "None":
                    cell.font = Font(color="FF0000", bold=True)
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

class IOCAnalyzerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Bulk IOC Analyzer - Developed by Naveen Kumar - Your Cyber Friend")
        self.root.geometry("800x600")
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(expand=True, fill='both', padx=10, pady=5)
        
        # Create tabs
        self.input_tab = ttk.Frame(self.notebook)
        self.output_tab = ttk.Frame(self.notebook)
        
        self.notebook.add(self.input_tab, text='Input')
        self.notebook.add(self.output_tab, text='Output')
        
        self.setup_input_tab()
        self.setup_output_tab()
        
        # API Key Entry
        self.api_frame = ttk.LabelFrame(self.input_tab, text="VirusTotal API Key")
        self.api_frame.pack(fill='x', padx=5, pady=5)
        self.api_key = tk.StringVar(value="")
        self.api_entry = ttk.Entry(self.api_frame, textvariable=self.api_key, show='*')
        self.api_entry.pack(fill='x', padx=5, pady=5)
        
        # Submit Button
        self.submit_btn = ttk.Button(self.input_tab, text="Start Analysis", command=self.start_analysis)
        self.submit_btn.pack(pady=10)
        
        # Progress Bar
        self.progress = ttk.Progressbar(self.input_tab, mode='indeterminate')
    
    def setup_input_tab(self):
        # Create frames for each input type
        input_types = ['Hashes', 'IP Addresses', 'Domains']
        self.text_widgets = {}
        
        for input_type in input_types:
            frame = ttk.LabelFrame(self.input_tab, text=input_type)
            frame.pack(fill='both', expand=True, padx=5, pady=5)
            
            text_widget = scrolledtext.ScrolledText(frame, height=5)
            text_widget.pack(fill='both', expand=True, padx=5, pady=5)
            self.text_widgets[input_type] = text_widget
    
    def setup_output_tab(self):
        self.output_text = scrolledtext.ScrolledText(self.output_tab)
        self.output_text.pack(fill='both', expand=True, padx=5, pady=5)
    
    def get_input_data(self):
        data = {}
        for input_type, widget in self.text_widgets.items():
            text = widget.get('1.0', 'end-1c')
            data[input_type] = [line.strip() for line in text.split('\n') if line.strip()]
        return data
    
    def start_analysis(self):
        # Disable submit button
        self.submit_btn.config(state='disabled')
        self.progress.pack(pady=5)
        self.progress.start()
        
        # Clear output
        self.output_text.delete('1.0', 'end')
        
        # Get input data
        input_data = self.get_input_data()
        
        # Start analysis in separate thread
        thread = threading.Thread(target=self.run_analysis, args=(input_data,))
        thread.daemon = True
        thread.start()
    
    def run_analysis(self, input_data):
        try:
            # Redirect print statements to the output text widget
            original_stdout = sys.stdout
            
            class StdoutRedirector:
                def __init__(self, widget):
                    self.widget = widget
                
                def write(self, string):
                    self.widget.after(0, self.widget.insert, 'end', string)
                    self.widget.after(0, self.widget.see, 'end')
                    self.widget.after(0, self.widget.update)
                
                def flush(self):
                    pass
            
            sys.stdout = StdoutRedirector(self.output_text)
            
            print("Starting analysis...")
            
            # Run the main analysis with the input data
            api_key = self.api_key.get()
            
            # Process data
            md5_df = process_md5(api_key, input_data['Hashes'])
            ips_df = process_ips(api_key, input_data['IP Addresses'])
            domains_df = process_urls(api_key, input_data['Domains'])
            
            # Save to Excel with formatting
            output_file = './IOC_Analysis_Results.xlsx'
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Write DataFrames
                md5_df.to_excel(writer, sheet_name='Hashes', index=False)
                ips_df.to_excel(writer, sheet_name='IPs', index=False)
                domains_df.to_excel(writer, sheet_name='Domains', index=False)
                
                # Apply formatting to each worksheet
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    format_worksheet(worksheet, sheet_name)
            
            def show_success():
                messagebox.showinfo("Success", f"Results have been saved to {output_file}")
            
            self.root.after(0, show_success)
            
        except Exception as error:
            def show_error(err=error):
                messagebox.showerror("Error", f"An error occurred: {str(err)}")
            self.root.after(0, show_error)
            
        finally:
            # Restore stdout
            sys.stdout = original_stdout
            
            def cleanup():
                self.submit_btn.config(state='normal')
                self.progress.stop()
                self.progress.pack_forget()
                
            self.root.after(0, cleanup)

def main():
    root = tk.Tk()
    app = IOCAnalyzerGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
